import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

event_path = os.getenv('GITHUB_EVENT_PATH')
with open(event_path, 'r') as f:
    event = json.load(f)
pr = event.get("pull_request")
if not pr:
    raise ValueError("No pull_request data found.")

req_id = pr.get("number")
title = pr.get("title")
author = pr.get("user", {}).get("login")
source_branch = pr.get("head", {}).get("ref")
target_branch = pr.get("base", {}).get("ref")
source_target = f"{source_branch} to {target_branch}"

action = "New Action"
comment = "New Comment"
merged_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

excel_file = "new_sheet.xlsx"
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Source Branch", "Author", "Action", "Comment", "Date"])

ws.append([source_branch, author, action, comment, merged_at])

wb.save(excel_file)
print(f"Logged new sheet")
