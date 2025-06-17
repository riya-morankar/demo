import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

event_path = os.getenv('GITHUB_EVENT_PATH')
with open(event_path, 'r') as f:
    event = json.load(f)
pr = event.get("pull_request")
if not pr:
    raise ValueError("No pull_request data found.")

req_id = pr.get("number")
title = pr.get("title")
author = pr.get("user", {}).get("login")
source_target = f"from {pr.get('head', {}).get('ref')} to {pr.get('base', {}).get('ref')}"

approvers_list = [r.get("login") for r in pr.get("requested_reviewers", [])]
approvers = ", ".join(approvers_list) if approvers_list else "N/A"

merged_at = pr.get("merged_at")
merged_date = merged_at[:10] if merged_at else datetime.now().strftime("%Y-%m-%d")

excel_file = "pull_requests.xlsx"
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Title", "Author", "Approvers", "Source / Target", "Date"])

ws.append([req_id, title, author, approvers, source_target, merged_date])

wb.save(excel_file)
print(f"Logged PR #{req_id}")
