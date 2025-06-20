import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import subprocess

event_path = os.getenv('GITHUB_EVENT_PATH')
with open(event_path, 'r') as f:
    event = json.load(f)
pr = event.get("pull_request")
if not pr:
    raise ValueError("No pull_request data found.")

req_id = pr.get("number")
title = pr.get("title")
author = pr.get("user", {}).get("login")
source_branch = pr.get("head", {}).get("ref")

if pr.merged:
    merge_commit_sha = pr.merge_commit_sha
    pr_commits = [commit.sha for commit in pr.get_commits()]
    if merge_commit_sha in pr_commits:
         action = "Squashed"
    else:
         action = "Merged"
else:
         action = "N/A"

# action = "Merged" if pr.get("merge_status") == "MERGED" else "Squashed"
# action = "Merged" if not pr.get("squash") else "Squashed"

comment = pr.get("body")
if not comment:
    comment="N/A"
merged_at = datetime.now().strftime("%Y-%m-%d")

merge_commit_sha = pr.get("merge_commit_sha")
if merge_commit_sha:
    change_id = subprocess.check_output(["git", "show", "-s", "--format=%H", merge_commit_sha]).decode("utf-8").strip()
else:
    change_id = "N/A"
    
excel_file = "new_sheet.xlsx"
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Source Branch", "Author", "Action", "Comment", "Date", "Change ID"])

ws.append([source_branch, author, action, comment, merged_at, change_id])
print(source_branch, author, action, comment, merged_at, change_id)
wb.save(excel_file)
print(f"Logged new sheet")
